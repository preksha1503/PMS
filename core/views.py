from datetime import timedelta

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models import Q
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import AssignProjectForm, LoginForm, ProjectForm, SignupForm
from .models import Notification, Project, ProjectAssignment, RoleRequest, User


def uses_director_ui(user):
    return user.is_superuser or user.has_management_role()


def normalize_status(value):
    normalized = (value or "").strip().lower()
    if "plan" in normalized:
        return "planning"
    if "complete" in normalized:
        return "completed"
    if "hold" in normalized:
        return "on_hold"
    if "progress" in normalized:
        return "in_progress"
    if "pending" in normalized:
        return "pending"
    return "pending"


def get_status_counts(projects):
    status_groups = projects.values("status").annotate(count=Count("id"))
    counts = {
        "planning": 0,
        "pending": 0,
        "completed": 0,
        "on_hold": 0,
        "in_progress": 0,
    }
    for item in status_groups:
        counts[normalize_status(item["status"])] += item["count"]
    return counts


def resolve_project_from_search(search_value):
    value = (search_value or "").strip()
    if not value:
        return None

    serial_part = value.split("-", 1)[0].strip()
    if serial_part.isdigit():
        project = Project.objects.filter(serial_number=int(serial_part)).first()
        if project:
            return project

    exact = Project.objects.filter(
        Q(initiative_scheme_project_portal__iexact=value)
        | Q(stakeholder_ministry__iexact=value)
        | Q(stakeholder_department__iexact=value)
        | Q(stakeholder_organization__iexact=value)
    ).first()
    if exact:
        return exact

    matches = Project.objects.filter(
        Q(initiative_scheme_project_portal__icontains=value)
        | Q(stakeholder_ministry__icontains=value)
        | Q(stakeholder_department__icontains=value)
        | Q(stakeholder_state__icontains=value)
        | Q(stakeholder_organization__icontains=value)
    ).order_by("serial_number")
    if matches.count() == 1:
        return matches.first()
    return None


def to_int_or_default(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def import_projects_from_excel(uploaded_file, created_by):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.datetime import from_excel
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is not installed.") from exc

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    imported = 0
    updated = 0
    skipped = 0

    from datetime import date as date_type
    from datetime import datetime

    def normalize_header(value):
        text = (str(value) if value is not None else "").strip().lower()
        if not text:
            return ""
        keep = []
        for ch in text:
            if ch.isalnum():
                keep.append(ch)
            elif ch in {" ", "/", "-", "_", "."}:
                keep.append(" ")
        return " ".join("".join(keep).split())

    header_aliases = {
        "serial_number": {"s no", "sno", "serial number", "serial", "project code", "projectcode", "code", "s no."},
        "website_url": {"website url", "website", "url", "websiteurl"},
        "stakeholder_ministry": {"stakeholder ministry", "ministry"},
        "stakeholder_department": {"stakeholder department", "department"},
        "stakeholder_state": {"stakeholder state", "state"},
        "stakeholder_state_ministry_department": {"stakeholder state ministry department", "state ministry department", "state ministry/department"},
        "stakeholder_organization": {"stakeholder organization", "organization"},
        "initiative_scheme_project_portal": {"initiative scheme project portal", "initiative", "project name", "project portal"},
        "genesis": {"genesis"},
        "year": {"year"},
        "project_manager_gis": {"project manager gis", "project manager - gis", "pm gis", "gis"},
        "project_manager_sw_mobi": {"project manager s w mobi", "project manager sw mobi", "project manager - s w mobi", "pm sw mobi", "sw mobi"},
        "start": {"start", "start date", "startdate"},
        "end": {"end", "end date", "enddate"},
        "status": {"status"},
    }

    def classify_header(cell_value):
        norm = normalize_header(cell_value)
        if not norm:
            return None
        for key, aliases in header_aliases.items():
            if norm in aliases:
                return key
        if "s no" in norm or "serial" in norm or "project code" in norm:
            return "serial_number"
        if "website" in norm and "url" in norm:
            return "website_url"
        if norm.startswith("website") or norm == "url":
            return "website_url"
        if "start" in norm and "date" in norm:
            return "start"
        if "end" in norm and "date" in norm:
            return "end"
        if norm == "project":
            return "initiative_scheme_project_portal"
        return None

    def as_date(value):
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date_type):
            return value
        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in (
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%d %b %Y",
                "%d %B %Y",
                "%b %d, %Y",
                "%B %d, %Y",
            ):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    # Discover header row/column positions (supports title/group rows above headers).
    header_row_index = None
    col_map = {}
    best_score = -1

    header_rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))

    def candidates_for(header_cells):
        candidates = {}
        for col_index, cell_value in enumerate(header_cells or []):
            key = classify_header(cell_value)
            if key and key not in candidates:
                candidates[key] = col_index
        return candidates

    for idx, row in enumerate(header_rows):
        # Single-row header
        cand = candidates_for(row)
        if {"serial_number", "website_url"}.issubset(set(cand)):
            score = len(cand)
            if score > best_score:
                best_score = score
                header_row_index = idx + 1
                col_map = cand

        # Two-row (group + subheader) header
        if idx + 1 < len(header_rows):
            next_row = header_rows[idx + 1] or ()
            max_len = max(len(row or ()), len(next_row))
            combined = []
            for col in range(max_len):
                top = (row[col] if row and col < len(row) else "") or ""
                bottom = (next_row[col] if col < len(next_row) else "") or ""
                combined.append(f"{top} {bottom}".strip())
            cand2 = candidates_for(combined)
            if {"serial_number", "website_url"}.issubset(set(cand2)):
                score = len(cand2)
                if score > best_score:
                    best_score = score
                    header_row_index = idx + 2
                    col_map = cand2

    if header_row_index is None:
        raise RuntimeError(
            "Could not find the header row. Please ensure the file contains 'S.No.' and 'Website URL' columns."
        )

    def get_cell(row, key, default=None):
        idx = col_map.get(key)
        if idx is None:
            return default
        if idx >= len(row):
            return default
        return row[idx]

    def add_text(defaults, field_name, key_name):
        if key_name not in col_map:
            return
        defaults[field_name] = (str(get_cell(row, key_name) or "")).strip()

    def add_int(defaults, field_name, key_name, fallback):
        if key_name not in col_map:
            return
        defaults[field_name] = to_int_or_default(get_cell(row, key_name), fallback)

    # Read rows after header.
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        serial_number = to_int_or_default(get_cell(row, "serial_number"), 0)
        website_url_value = get_cell(row, "website_url")
        website_url = (str(website_url_value).strip() if website_url_value is not None else "")
        if not serial_number or not website_url:
            skipped += 1
            continue

        start_date = as_date(get_cell(row, "start"))
        end_date = as_date(get_cell(row, "end"))
        defaults = {"website_url": website_url, "created_by": created_by}
        add_text(defaults, "stakeholder_ministry", "stakeholder_ministry")
        add_text(defaults, "stakeholder_department", "stakeholder_department")
        add_text(defaults, "stakeholder_state", "stakeholder_state")
        add_text(defaults, "stakeholder_state_ministry_department", "stakeholder_state_ministry_department")
        add_text(defaults, "stakeholder_organization", "stakeholder_organization")
        add_text(defaults, "initiative_scheme_project_portal", "initiative_scheme_project_portal")
        add_text(defaults, "genesis", "genesis")
        defaults["year"] = to_int_or_default(get_cell(row, "year"), timezone.now().year)
        add_text(defaults, "project_manager_gis", "project_manager_gis")
        add_text(defaults, "project_manager_sw_mobi", "project_manager_sw_mobi")
        defaults["start"] = start_date or timezone.now().date()
        defaults["end"] = end_date or timezone.now().date()
        if "status" in col_map:
            defaults["status"] = (str(get_cell(row, "status") or "Pending")).strip() or "Pending"
        else:
            defaults["status"] = "Pending"
        if not (defaults.get("initiative_scheme_project_portal") or "").strip():
            defaults["initiative_scheme_project_portal"] = f"Project {serial_number}"
        _, was_created = Project.objects.update_or_create(
            serial_number=serial_number,
            defaults=defaults,
        )
        if was_created:
            imported += 1
        else:
            updated += 1

    if imported == 0 and updated == 0 and skipped:
        raise RuntimeError(
            "No rows were imported. Please ensure each row has 'S.No.' (or Project Code) and 'Website URL' filled."
        )

    return imported, updated


def signup_view(request):
    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect("role_request")
    else:
        form = SignupForm()
    return render(request, "core/signup.html", {"form": form})


def login_view(request):
    form = LoginForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        login(request, form.get_user())
        return redirect("dashboard")
    return render(request, "core/login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def dashboard_view(request):
    if request.user.is_superuser or request.user.role == User.Role.SUPER_ADMIN:
        return redirect("super_admin_dashboard")
    if request.user.role == User.Role.PENDING:
        return redirect("role_request")
    if not request.user.has_management_role() and not request.user.is_superuser:
        return redirect("view_projects")
    projects = Project.objects.all()
    status_counts = get_status_counts(projects)
    total_projects = projects.count()
    total_users = User.objects.count()
    month_label = timezone.now().strftime("%b %Y")
    today = timezone.localdate()

    overdue_projects = list(Project.objects.filter(end__lte=today).order_by("end", "serial_number"))
    overdue_projects = [p for p in overdue_projects if normalize_status(p.status) != "completed"]
    overdue_count = len(overdue_projects)
    overdue_items = overdue_projects[:5]
    overdue_more = max(0, overdue_count - len(overdue_items))

    def percent(count):
        if total_projects == 0:
            return 0
        return round((count / total_projects) * 100, 2)

    segments = [
        ("pending", "#3f7ae0", percent(status_counts["pending"])),
        ("in_progress", "#f2c01f", percent(status_counts["in_progress"])),
        ("completed", "#5fcb85", percent(status_counts["completed"])),
        ("on_hold", "#ec6e6e", percent(status_counts["on_hold"])),
        ("planning", "rgba(139,92,246,0.6)", percent(status_counts["planning"])),
    ]
    if total_projects == 0:
        pie_style = ""
    else:
        cursor = 0.0
        parts = []
        for _, color, pct in segments:
            end = cursor + pct
            parts.append(f"{color} {cursor:.2f}% {end:.2f}%")
            cursor = end
        pie_style = f"background: conic-gradient({', '.join(parts)});"

    max_count = max(status_counts.values()) if total_projects else 0
    if max_count == 0:
        bar_heights = {key: 0 for key in status_counts}
    else:
        bar_heights = {
            key: round((value / max_count) * 100, 1)
            for key, value in status_counts.items()
        }

    return render(
        request,
        "core/director_dashboard.html",
        {
            "status_counts": status_counts,
            "total_projects": total_projects,
            "total_users": total_users,
            "pie_style": pie_style,
            "pie_empty": total_projects == 0,
            "bar_heights": bar_heights,
            "month_label": month_label,
            "overdue_projects": overdue_projects,
            "overdue_count": overdue_count,
            "overdue_items": overdue_items,
            "overdue_more": overdue_more,
            "overdue_today": today,
        },
    )


@login_required
def super_admin_dashboard_view(request):
    if not request.user.is_superuser and request.user.role != User.Role.SUPER_ADMIN:
        return redirect("dashboard")

    projects = Project.objects.all()
    status_counts = get_status_counts(projects)
    total_projects = projects.count()
    total_users = User.objects.count()
    total_pending_requests = RoleRequest.objects.filter(status=RoleRequest.Status.PENDING).count()
    month_label = timezone.now().strftime("%b %Y")

    def percent(count):
        if total_projects == 0:
            return 0
        return round((count / total_projects) * 100, 2)

    segments = [
        ("pending", "#3f7ae0", percent(status_counts["pending"])),
        ("in_progress", "#f2c01f", percent(status_counts["in_progress"])),
        ("completed", "#5fcb85", percent(status_counts["completed"])),
        ("on_hold", "#ec6e6e", percent(status_counts["on_hold"])),
        ("planning", "rgba(139,92,246,0.6)", percent(status_counts["planning"])),
    ]
    if total_projects == 0:
        pie_style = ""
    else:
        cursor = 0.0
        parts = []
        for _, color, pct in segments:
            end = cursor + pct
            parts.append(f"{color} {cursor:.2f}% {end:.2f}%")
            cursor = end
        pie_style = f"background: conic-gradient({', '.join(parts)});"

    max_count = max(status_counts.values()) if total_projects else 0
    if max_count == 0:
        bar_heights = {key: 0 for key in status_counts}
    else:
        bar_heights = {
            key: round((value / max_count) * 100, 1)
            for key, value in status_counts.items()
        }

    temp_credentials = None
    if request.method == "POST" and request.POST.get("action") == "generate_temp_admin":
        base = timezone.now().strftime("%Y%m%d%H%M%S")
        username = f"temp_super_admin_{base}"
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"temp_super_admin_{base}_{counter}"
            counter += 1
        password = User.objects.make_random_password(length=12)
        temp_user = User.objects.create_user(username=username, password=password, role=User.Role.SUPER_ADMIN)
        temp_user.is_staff = True
        temp_user.is_superuser = True
        temp_user.save(update_fields=["is_staff", "is_superuser"])
        temp_credentials = {"username": username, "password": password}
        messages.success(request, "Temporary super admin created. Copy the credentials now.")

    page_links = [
        {"label": "Dashboard", "url": "dashboard"},
        {"label": "User Management", "url": "super_admin_users"},
        {"label": "Add Project", "url": "project_create"},
        {"label": "View Projects", "url": "view_projects"},
        {"label": "Assign Project", "url": "assign_project"},
        {"label": "Assign To", "url": "assign_to"},
        {"label": "Excel Entry", "url": "excel_entry"},
        {"label": "Role Access", "url": "role_access"},
        {"label": "Profile", "url": "profile"},
        {"label": "Change Password", "url": "change_password"},
        {"label": "Help", "url": "help_page"},
        {"label": "Contact Us", "url": "contact_us"},
        {"label": "Chat", "url": "chat_page"},
        {"label": "Notifications", "url": "notifications_page"},
    ]

    return render(
        request,
        "core/super_admin_dashboard.html",
        {
            "status_counts": status_counts,
            "total_projects": total_projects,
            "total_users": total_users,
            "pie_style": pie_style,
            "pie_empty": total_projects == 0,
            "bar_heights": bar_heights,
            "month_label": month_label,
            "total_pending_requests": total_pending_requests,
            "temp_credentials": temp_credentials,
            "page_links": page_links,
        },
    )


@login_required
def super_admin_users_view(request):
    if not request.user.is_superuser and request.user.role != User.Role.SUPER_ADMIN:
        return redirect("dashboard")

    q = (request.GET.get("q") or "").strip()
    role_filter = (request.GET.get("role") or "").strip()
    users = User.objects.all().order_by("username")
    if q:
        users = users.filter(
            Q(username__icontains=q)
            | Q(email__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        )
    if role_filter:
        users = users.filter(role=role_filter)

    if request.method == "POST":
        action = request.POST.get("action")
        target_id = request.POST.get("user_id")
        target = User.objects.filter(id=target_id).first()
        if not target:
            messages.error(request, "User not found.")
            return redirect("super_admin_users")

        if action == "set_role":
            role_value = (request.POST.get("role") or "").strip()
            allowed_values = {value for value, _ in User.Role.choices}
            confirm_super_admin = request.POST.get("confirm_super_admin") == "yes"
            if role_value not in allowed_values:
                messages.error(request, "Please select a valid role.")
            elif role_value == User.Role.SUPER_ADMIN and not confirm_super_admin:
                messages.error(request, "Confirm Super Admin assignment before updating the role.")
            else:
                target.role = role_value
                target.save(update_fields=["role"])
                messages.success(request, f"Updated role for {target.username}.")
        elif action == "toggle_active":
            if target.id == request.user.id:
                messages.error(request, "You cannot deactivate your own account.")
            else:
                target.is_active = not target.is_active
                target.save(update_fields=["is_active"])
                state = "activated" if target.is_active else "deactivated"
                messages.success(request, f"{target.username} has been {state}.")
        else:
            messages.error(request, "Invalid action.")

        return redirect("super_admin_users")

    return render(
        request,
        "core/super_admin_users.html",
        {
            "users": users,
            "q": q,
            "role_filter": role_filter,
            "role_choices": User.Role.choices,
        },
    )


@login_required
def project_create_view(request):
    if not request.user.can_manage_projects():
        return redirect("dashboard")
    if request.method == "POST":
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            return redirect("dashboard")
    else:
        form = ProjectForm()
    template_name = "core/director_project_form.html" if uses_director_ui(request.user) else "core/project_form.html"
    return render(request, template_name, {"form": form, "mode": "Create"})


@login_required
def project_edit_view(request, project_id):
    if not request.user.can_manage_projects():
        return redirect("dashboard")
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            return redirect("dashboard")
    else:
        form = ProjectForm(instance=project)
    template_name = "core/director_project_form.html" if uses_director_ui(request.user) else "core/project_form.html"
    return render(request, template_name, {"form": form, "mode": "Edit"})


@login_required
def project_detail_view(request, project_id):
    if request.user.role == User.Role.PENDING:
        return redirect("role_request")
    project = get_object_or_404(Project, id=project_id)
    return render(request, "core/director_project_detail.html", {"project": project})


@login_required
def project_delete_view(request, project_id):
    if not request.user.can_manage_projects():
        return redirect("dashboard")
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        project.delete()
        messages.success(request, "Project deleted successfully.")
    return redirect("view_projects")


@login_required
def assign_project_view(request):
    return _assign_page_handler(request, active_page="assign_project", page_title="Assign Project")


@login_required
def assign_to_view(request):
    q = (request.GET.get("q") or "").strip()
    assignments = ProjectAssignment.objects.select_related("project", "assigned_to", "assigned_by")
    if not request.user.can_manage_projects():
        assignments = assignments.filter(assigned_to=request.user)
    if q:
        assignments = assignments.filter(
            Q(project__initiative_scheme_project_portal__icontains=q)
            | Q(project__stakeholder_organization__icontains=q)
            | Q(project__status__icontains=q)
            | Q(assigned_to__username__icontains=q)
            | Q(assigned_by__username__icontains=q)
            | Q(assigned_role__icontains=q)
        )
        if q.isdigit():
            assignments = assignments | ProjectAssignment.objects.select_related(
                "project", "assigned_to", "assigned_by"
            ).filter(project__serial_number=int(q))
    assignments = assignments.order_by("-assigned_at").distinct()
    return render(
        request,
        "core/director_assign_to.html",
        {
            "assignments": assignments,
            "q": q,
            "total_assignments": assignments.count(),
        },
    )


def _assign_page_handler(request, active_page, page_title):
    if not request.user.can_manage_projects():
        return redirect("dashboard")

    if request.method == "POST":
        form = AssignProjectForm(request.POST)
        if form.is_valid():
            project = resolve_project_from_search(form.cleaned_data["search_project"])
            if not project:
                form.add_error("search_project", "Project not found. Use exact serial number or select from suggestions.")
            else:
                assignee_email = (form.cleaned_data.get("assigned_to") or "").strip().lower()
                assignee = User.objects.filter(email__iexact=assignee_email).first()
                if not assignee:
                    form.add_error("assigned_to", "No user found with this email.")
                else:
                    ProjectAssignment.objects.update_or_create(
                        project=project,
                        assigned_to=assignee,
                        defaults={
                            "assigned_by": request.user,
                            "notes": form.cleaned_data["notes"],
                            "assigned_role": "",
                        },
                    )
                    actor_label = request.user.get_role_display() if hasattr(request.user, "get_role_display") else "Admin"
                    Notification.objects.create(
                        user=assignee,
                        project=project,
                        created_by=request.user,
                        message=f"{actor_label} assigned you a new project: {project.initiative_scheme_project_portal}",
                    )
                    messages.success(request, "Project assigned successfully.")
                    return redirect(active_page)
    else:
        form = AssignProjectForm()

    project_suggestions = Project.objects.order_by("serial_number").values(
        "serial_number",
        "initiative_scheme_project_portal",
        "stakeholder_organization",
    )
    user_suggestions = User.objects.filter(is_active=True).exclude(email="").order_by("email").values(
        "email",
        "role",
        "username",
    )
    return render(
        request,
        "core/director_assign_project.html",
        {
            "form": form,
            "project_suggestions": project_suggestions,
            "user_suggestions": user_suggestions,
            "active_page": active_page,
            "page_title": page_title,
        },
    )


@login_required
def view_projects_view(request):
    if request.user.role == User.Role.PENDING:
        return redirect("role_request")

    projects = Project.objects.all()
    if request.user.role in {User.Role.DEVELOPER, User.Role.ASSISTANT_PROJECT_MANAGER} and not request.user.is_superuser:
        assignment_filter = Q(assignments__assigned_to=request.user) | Q(assignments__assigned_role=request.user.role)
        assigned_by_super_admin = Q(assignments__assigned_by__role=User.Role.SUPER_ADMIN) | Q(
            assignments__assigned_by__is_superuser=True
        )
        projects = Project.objects.filter(
            assignment_filter & assigned_by_super_admin
        ).distinct()

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    today = timezone.localdate()

    filtered_projects = projects
    if q:
        query_filter = (
            Q(initiative_scheme_project_portal__icontains=q)
            | Q(stakeholder_ministry__icontains=q)
            | Q(stakeholder_department__icontains=q)
            | Q(stakeholder_state__icontains=q)
            | Q(stakeholder_organization__icontains=q)
            | Q(project_manager_gis__icontains=q)
            | Q(project_manager_sw_mobi__icontains=q)
        )
        if q.isdigit():
            query_filter = query_filter | Q(serial_number=int(q))
        filtered_projects = filtered_projects.filter(query_filter)
    if status_filter:
        filtered_projects = [p for p in filtered_projects if normalize_status(p.status) == status_filter]
    else:
        filtered_projects = list(filtered_projects)

    status_counts = get_status_counts(projects)

    reminder_window_days = 7
    due_soon_cutoff = today + timedelta(days=reminder_window_days)
    due_soon_projects = list(projects.filter(end__gte=today, end__lte=due_soon_cutoff).order_by("end", "serial_number"))
    due_soon_projects = [p for p in due_soon_projects if normalize_status(p.status) != "completed"]
    due_soon_count = len(due_soon_projects)
    due_soon_items = due_soon_projects[:5]
    due_soon_more = max(0, due_soon_count - len(due_soon_items))
    due_soon_primary = due_soon_items[0] if due_soon_items else None
    return render(
        request,
        "core/director_view_projects.html",
        {
            "projects": filtered_projects,
            "total_projects": projects.count(),
            "showing_count": len(filtered_projects),
            "status_counts": status_counts,
            "q": q,
            "status_filter": status_filter,
            "can_manage": request.user.can_manage_projects(),
            "due_soon_projects": due_soon_projects,
            "due_soon_count": due_soon_count,
            "due_soon_items": due_soon_items,
            "due_soon_more": due_soon_more,
            "due_soon_primary": due_soon_primary,
            "due_soon_today": today,
            "due_soon_cutoff": due_soon_cutoff,
            "due_soon_window_days": reminder_window_days,
        },
    )


@login_required
def projects_download_excel_view(request):
    if request.user.role == User.Role.PENDING:
        return redirect("role_request")

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    projects = Project.objects.all()
    if request.user.role in {User.Role.DEVELOPER, User.Role.ASSISTANT_PROJECT_MANAGER} and not request.user.is_superuser:
        assignment_filter = Q(assignments__assigned_to=request.user) | Q(assignments__assigned_role=request.user.role)
        assigned_by_super_admin = Q(assignments__assigned_by__role=User.Role.SUPER_ADMIN) | Q(
            assignments__assigned_by__is_superuser=True
        )
        projects = Project.objects.filter(
            assignment_filter & assigned_by_super_admin
        ).distinct()
    filtered_projects = projects
    if q:
        query_filter = (
            Q(initiative_scheme_project_portal__icontains=q)
            | Q(stakeholder_ministry__icontains=q)
            | Q(stakeholder_department__icontains=q)
            | Q(stakeholder_state__icontains=q)
            | Q(stakeholder_organization__icontains=q)
            | Q(project_manager_gis__icontains=q)
            | Q(project_manager_sw_mobi__icontains=q)
        )
        if q.isdigit():
            query_filter = query_filter | Q(serial_number=int(q))
        filtered_projects = filtered_projects.filter(query_filter)

    filtered_projects = list(filtered_projects)
    if status_filter:
        filtered_projects = [p for p in filtered_projects if normalize_status(p.status) == status_filter]

    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return HttpResponse(
            "openpyxl is not installed. Install it with: pip install openpyxl",
            status=500,
            content_type="text/plain",
        )

    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    ws.append(["Projects Export"])
    ws.append(
        [
            "S.No.",
            "Website URL",
            "Stakeholder - Ministry",
            "Stakeholder - Department",
            "Stakeholder - State",
            "Stakeholder - State Ministry/Department",
            "Stakeholder - Organization",
            "Initiative/Scheme/Project Portal",
            "Genesis",
            "Year",
            "Project Manager - GIS",
            "Project Manager - S/W & Mobi",
            "Start",
            "End",
            "Status",
        ]
    )

    for project in filtered_projects:
        ws.append(
            [
                project.serial_number,
                project.website_url,
                project.stakeholder_ministry,
                project.stakeholder_department,
                project.stakeholder_state,
                project.stakeholder_state_ministry_department,
                project.stakeholder_organization,
                project.initiative_scheme_project_portal,
                project.genesis,
                project.year,
                project.project_manager_gis,
                project.project_manager_sw_mobi,
                project.start,
                project.end,
                project.status,
            ]
        )

    ws.freeze_panes = "A3"

    date_format = "DD-MMM-YYYY"
    for row in ws.iter_rows(min_row=3, min_col=13, max_col=14):
        for cell in row:
            if cell.value:
                cell.number_format = date_format

    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    filename = f"projects_{timestamp}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def excel_entry_view(request):
    if not request.user.can_manage_projects():
        return redirect("dashboard")

    if request.method == "POST":
        upload = request.FILES.get("excel_file")
        if not upload:
            messages.error(request, "Please choose an Excel file.")
        elif not upload.name.lower().endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported. Please save your file as .xlsx.")
        else:
            try:
                imported, updated = import_projects_from_excel(upload, request.user)
                messages.success(request, f"Import complete. Added {imported} and updated {updated} projects.")
                return redirect("view_projects")
            except RuntimeError as exc:
                message = str(exc)
                if "openpyxl" in message.lower() and "not installed" in message.lower():
                    message = f"{message} Install it with: pip install openpyxl"
                messages.error(request, message)
            except Exception:
                messages.error(request, "Could not import this file. Please verify the sheet format.")

    return render(request, "core/director_excel_entry.html")


@login_required
def profile_view(request):
    user = request.user
    context = {
        "profile_name": user.get_full_name() or user.username,
        "profile_role": user.get_role_display() if hasattr(user, "get_role_display") else "N/A",
        "email": user.email or "N/A",
        "department": "N/A",
        "employee_id": "N/A",
        "phone": "N/A",
    }
    return render(request, "core/director_profile.html", context)


@login_required
def role_request_view(request):
    if request.user.role == User.Role.SUPER_ADMIN or request.user.is_superuser:
        return redirect("super_admin_dashboard")

    allowed_roles = [
        choice for choice in User.Role.choices if choice[0] not in (User.Role.SUPER_ADMIN, User.Role.PENDING)
    ]
    pending_request = RoleRequest.objects.filter(user=request.user, status=RoleRequest.Status.PENDING).first()
    latest_request = RoleRequest.objects.filter(user=request.user).order_by("-created_at").first()

    if request.method == "POST":
        requested_role = (request.POST.get("requested_role") or "").strip()
        allowed_values = {value for value, _ in allowed_roles}
        if requested_role not in allowed_values:
            messages.error(request, "Please select a valid role.")
        else:
            if pending_request:
                pending_request.requested_role = requested_role
                pending_request.save(update_fields=["requested_role", "updated_at"])
            else:
                RoleRequest.objects.create(user=request.user, requested_role=requested_role)
            messages.success(request, "Role request submitted for approval.")
            return redirect("role_request")

    return render(
        request,
        "core/director_role_request.html",
        {
            "allowed_roles": allowed_roles,
            "pending_request": pending_request,
            "latest_request": latest_request,
        },
    )


@login_required
def role_access_view(request):
    if request.user.role != User.Role.SUPER_ADMIN and not request.user.is_superuser:
        return redirect("dashboard")

    pending_requests = RoleRequest.objects.filter(status=RoleRequest.Status.PENDING).select_related("user")
    allowed_roles = [
        choice for choice in User.Role.choices if choice[0] not in (User.Role.SUPER_ADMIN, User.Role.PENDING)
    ]

    if request.method == "POST":
        request_id = request.POST.get("request_id")
        action = request.POST.get("action")
        role_value = request.POST.get("role")
        role_request = RoleRequest.objects.filter(id=request_id).select_related("user").first()
        if not role_request or role_request.status != RoleRequest.Status.PENDING:
            messages.error(request, "Role request not found or already processed.")
            return redirect("role_access")

        if action == "approve":
            allowed_values = {value for value, _ in allowed_roles}
            if role_value not in allowed_values:
                messages.error(request, "Please select a valid role.")
                return redirect("role_access")
            role_request.user.role = role_value
            role_request.user.save(update_fields=["role"])
            role_request.status = RoleRequest.Status.APPROVED
            role_request.reviewed_by = request.user
            role_request.reviewed_at = timezone.now()
            role_request.save(update_fields=["status", "reviewed_by", "reviewed_at", "updated_at"])
            messages.success(request, f"Approved role for {role_request.user.username}.")
        elif action == "reject":
            role_request.status = RoleRequest.Status.REJECTED
            role_request.reviewed_by = request.user
            role_request.reviewed_at = timezone.now()
            role_request.save(update_fields=["status", "reviewed_by", "reviewed_at", "updated_at"])
            messages.success(request, f"Rejected role request for {role_request.user.username}.")
        else:
            messages.error(request, "Invalid action.")
        return redirect("role_access")

    return render(
        request,
        "core/director_role_access.html",
        {
            "pending_requests": pending_requests,
            "allowed_roles": allowed_roles,
        },
    )


class DirectorPasswordChangeView(PasswordChangeView):
    template_name = "core/director_change_password.html"
    success_url = reverse_lazy("profile")

    def form_valid(self, form):
        messages.success(self.request, "Password updated successfully.")
        return super().form_valid(form)


@login_required
def help_view(request):
    return render(request, "core/director_help.html")


@login_required
def contact_us_view(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        message = (request.POST.get("message") or "").strip()
        if not name or not email or not message:
            messages.error(request, "Please fill all fields before submitting.")
        else:
            messages.success(request, "Your message has been submitted successfully.")
            return redirect("contact_us")
    return render(request, "core/director_contact_us.html")


@login_required
def chat_view(request):
    return render(request, "core/director_chat.html")


@login_required
def notifications_view(request):
    notifications = Notification.objects.filter(user=request.user).select_related("project", "created_by")
    return render(request, "core/director_notifications.html", {"notifications": notifications})


@login_required
@require_POST
def notification_read_view(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save(update_fields=["is_read"])
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({"ok": True, "unread_count": unread_count})


@login_required
@require_POST
def notifications_read_all_view(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"ok": True, "unread_count": 0})


